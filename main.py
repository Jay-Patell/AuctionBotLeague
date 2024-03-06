import os

import discord
from discord.ext import commands, tasks
from discord.ui import Button, View
import asyncio
import json
import openpyxl
from openpyxl import load_workbook
import pprint


#############################################################
# Custom help:
#############################################################
class CustomHelp(commands.HelpCommand):
    def __init__(self):
        super().__init__()

    async def send_bot_help(self, mapping):
      channel = self.get_destination()
      embed = discord.Embed(title="Help", description="List of available commands:", color=discord.Color.blue())
      for cog, commands in mapping.items():
        filtered = await self.filter_commands(commands, sort=True)
        if filtered:
          cog_name = 'No Category' if cog is None else cog.qualified_name
          command_listings = [f'`{self.context.clean_prefix}{command.name}`: {command.short_doc}' for command in filtered]
          embed.add_field(name=cog_name, value="\n".join(command_listings), inline=False)
      await channel.send(embed=embed)


    async def send_command_help(self, command):
        embed = discord.Embed(title=self.get_command_signature(command),
                              description=command.help or "No description provided.",
                              color=discord.Color.green())
        await self.get_destination().send(embed=embed)

    def get_command_signature(self, command):
        return f'{self.context.clean_prefix}{command.qualified_name} {command.signature}'

    async def send_group_help(self, group):
        embed = discord.Embed(title=self.get_command_signature(group),
                              description=group.help or "No description provided.",
                              color=discord.Color.orange())
        if group.commands:
            embed.add_field(name="Subcommands", value="\n".join([f'{command.name} - {command.short_doc}' for command in group.commands]), inline=False)
        await self.get_destination().send(embed=embed)




#############################################################
# Declarations:
#############################################################


intents = discord.Intents.default()
intents.messages = True
intents.message_content = True

bot = commands.Bot(command_prefix='!', intents=intents)
bot.help_command = CustomHelp()

TEAMS = {}  
PLAYERS_FOR_AUCTION = []  
UNSOLD_PLAYERS = []
MAX_TEAM_SIZE = 18
USERS = {}  }}
MAX_PURSE = 20000000
AUCTION_CHANNEL = None
auctioneer_id = 256972361918578688  
AUTHORIZED_USER_IDS = {'256972361918578688', '1111497896018313268'}




#############################################################
# Data Management:
#############################################################
def save_data():
  # This function saves the current state of players and teams to a file
  data = {
      "players_for_auction": PLAYERS_FOR_AUCTION,
      "teams": TEAMS,
      "unsold": UNSOLD_PLAYERS,
      "users": USERS
  }
  with open('auction_data.json', 'w') as f:
      json.dump(data, f, indent=4)

def load_data():
  global PLAYERS_FOR_AUCTION, TEAMS, USERS

  try:
      with open('auction_data.json', 'r') as f:
          data = json.load(f)
          PLAYERS_FOR_AUCTION = data.get("players_for_auction", [])
          TEAMS = data.get("teams", {})
          UNSOLD_PLAYERS = data.get('unsold', [])
          USERS = data.get("users", {})
  except FileNotFoundError:
      PLAYERS_FOR_AUCTION = []
      TEAMS = {}
      USERS = {}

load_data()

@bot.command(name='set_purse', help='Sets the purse amount for a specified Discord user.\nUsage: !set_purse [@user] [amount]\nExample: !set_purse @JohnDoe 50000')
@commands.has_permissions(administrator=True)  # Ensure only admins can set purses
async def set_purse(ctx, user: discord.User, amount: int):
    USERS[user.id] = {"purse": amount}
    save_data()  

    await ctx.send(f"Set the purse for {user.display_name} to {amount}.")

#############################################################
# Auction Channel:
#############################################################
@bot.command(name='set_auction_channel')
@commands.has_permissions(administrator=True)  # Ensure only users with administrator permissions can set the channel
async def set_auction_channel(ctx):
    global AUCTION_CHANNEL
    AUCTION_CHANNEL = ctx.channel.id  # Set the current channel as the auction channel
    await ctx.send(f"Auction channel set to {ctx.channel.name}")

#############################################################
# team logics:
#############################################################
@bot.command(name='delete_team')
@commands.has_permissions(administrator=True)  
async def delete_team(ctx, team_name: str):
    # Check if the team exists
    if team_name not in TEAMS:
        await ctx.send(f"Team '{team_name}' does not exist.")
        return

    # Optional: Check if the requester is the team owner or an admin
    if TEAMS[team_name]['owner'] != ctx.author.id and not ctx.author.guild_permissions.administrator:
        await ctx.send("You do not have permission to delete this team.")
        return

    # Delete the team
    del TEAMS[team_name]
    save_data() 

    await ctx.send(f"Team '{team_name}' has been successfully deleted.")


@bot.command(name='create_team', help='Creates a new team with a specified name, maximum size, and purse.\nUsage: !create_team [team_name] [max_size] [purse]\nExample: !create_team "Dream Team" 11 100000')
async def create_team(ctx, team_name: str, max_size: int = MAX_TEAM_SIZE, purse: int = MAX_PURSE):
  # Check if the team name already exists to prevent duplicates
  if team_name in TEAMS:
      await ctx.send(f"Team '{team_name}' already exists.")
      return

  # Correctly obtain the user's ID who issued the command
  owner_id = ctx.author.id 

  # Create the team and associate it with the user's ID
  TEAMS[team_name] = {
      'owner': owner_id,  
      'max_size': max_size,
      'purse': purse,
      'players': []
  }
  if owner_id not in USERS:
    USERS[owner_id] = {'purse': purse}
  save_data()  
  pprint.pprint(USERS)
  await ctx.send(f"Team {team_name} created with max size {max_size}, purse {purse}, and owner {ctx.author.display_name}.")

@bot.command(name='add_player_to_team', help='Adds a specified player to a specified team.\nUsage: !add_player_to_team [team_name] [player_name]\nExample: !add_player_to_team "Dream Team" "John Doe"')
async def add_player_to_team(ctx, team_name: str, *, player_name: str):
    if team_name not in TEAMS:
        await ctx.send("Team does not exist.")
        return
    if len(TEAMS[team_name]["players"]) >= TEAMS[team_name]["max_size"]:
        await ctx.send("Team is at maximum capacity.")
        return
    TEAMS[team_name]["players"].append(player_name)
    save_data()
    await ctx.send(f"Added {player_name} to {team_name}.")

@bot.command(name='remove_player', help='Removes a specified player from a specified team.\nUsage: !remove_player [team_name] [player_name]\nExample: !remove_player "Dream Team" "John Doe"')
async def remove_player(ctx, team_name: str, *, player_name: str):
  try:
    # Check if the team exists
    if team_name not in TEAMS:
        await ctx.send(f"Team '{team_name}' not found.")
        return

    # Check if the player is in the team
    if player_name not in TEAMS[team_name]['players']:
        await ctx.send(f"Player '{player_name}' is not in Team '{team_name}'.")
        return

    # Log current state before removal for debugging
    print(f"Before removal: {TEAMS[team_name]['players']}")

    # Remove the player from the team
    TEAMS[team_name]['players'].remove(player_name)
    save_data()  # Save the updated data to JSON

    # Log current state after removal for debugging
    print(f"After removal: {TEAMS[team_name]['players']}")

    await ctx.send(f"Removed player '{player_name}' from Team '{team_name}'.")
  except Exception as e:
    print(f"Error occurred: {e}")



@bot.command(name='team_info', help='Displays information about a specified team, including players, max size, and remaining purse.\nUsage: !team_info [team_name]\nExample: !team_info "Dream Team"')
async def team_info(ctx, team_name: str):
          # Check if the team exists
          if team_name in TEAMS:
              team = TEAMS[team_name]
              players = team.get("players", [])
              max_size = team.get("max_size", 0)
              purse = team.get("purse", 0)
              owner_id = team.get("owner")

              # Format player names as a list
              if players:
                  player_list = '\n'.join([f'- {player}' for player in players])
              else:
                  player_list = 'No players'

              owner = await ctx.bot.fetch_user(owner_id)
              owner_name = owner.display_name if owner else "Unknown"

              # purse format
              formatted_purse = f"${purse:,.2f}"

              # Prepare the embed message
              embed = discord.Embed(title=f"Team: {team_name}", color=discord.Color.blue())
              embed.add_field(name="Owner", value=owner_name, inline=False)
              embed.add_field(name="Max Size", value=str(max_size), inline=True)
              embed.add_field(name="Purse", value=formatted_purse, inline=True)
              embed.add_field(name="Players", value=player_list, inline=False)

              await ctx.send(embed=embed)
              pprint.pprint(USERS)
          else:
              await ctx.send(f"Team '{team_name}' not found.")



#############################################################
# player logics:
#############################################################
@bot.command(name='add_player_for_auction', help='Adds a player to the auction list with a specified category, first name, last name, and base price.\nUsage: !add_player_for_auction [category] [first_name] [last_name] [base_price]\nExample: !add_player_for_auction "Batsman" "John" "Doe" 50000')
async def add_player_for_auction(ctx, category: str, first_name: str, last_name: str, base_price: int):
    if category not in ["Batsmen", "Allrounders", "Bowlers"]:
        await ctx.send("Invalid category. Choose from Batsmen, Allrounders, Bowlers.")
        return
    player = {"name": f"{first_name} {last_name}", "category": category, "base_price": base_price}
    PLAYERS_FOR_AUCTION.append(player)
    save_data()
    await ctx.send(f"Added {player['name']} to the auction list.")

@bot.command(name='load_players_from_excel')
async def load_players_from_excel(ctx, file_path: str):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        category, first_name, last_name, base_price = row
        PLAYERS_FOR_AUCTION.append({"name": f"{first_name} {last_name}", "category": category, "base_price": base_price})
    await ctx.send("Players loaded from Excel.")

@bot.command(name='remove_player_from_auction', help='Removes a specified player from the auction list or removes all players if specified.\nUsage: !remove_player_from_auction [player_name | ALL]\nExample: !remove_player_from_auction "John Doe", !remove_player_from_auction ALL')
async def remove_player_from_auction(ctx, player_name: str):
    global PLAYERS_FOR_AUCTION  

    # Check if we're removing all players
    if player_name.upper() == 'ALL':
        PLAYERS_FOR_AUCTION = []  # Clear the list
        await ctx.send("All players have been removed from the auction.")
        save_data()  
        return

    # Find and remove the specified player
    for player in PLAYERS_FOR_AUCTION:
        if player_name == player['name']:
            PLAYERS_FOR_AUCTION.remove(player)
            await ctx.send(f"Removed {player_name} from the auction.")
            save_data()  
            return

    # If the player was not found
    await ctx.send(f"Player {player_name} not found in the auction.")

@bot.command(name='view_auction_players', help='Displays the list of players currently available for auction.')
async def view_auction_players(ctx):
    if not PLAYERS_FOR_AUCTION:
        await ctx.send("There are currently no players listed for auction.")
        return

    # Constructing the message
    players_list = []
    for player in PLAYERS_FOR_AUCTION:
        player_info = f"**Name**: {player['name']}, **Category**: {player['category']}, **Base Price**: {player['base_price']}"
        players_list.append(player_info)

    # Splitting the message if it's too long for one Discord message
    message_chunks = [players_list[i:i + 10] for i in range(0, len(players_list), 10)]  # Split list into chunks of 10

    for chunk in message_chunks:
        embed = discord.Embed(title="Auction Players List", description="\n".join(chunk), color=discord.Color.blue())
        await ctx.send(embed=embed)

@bot.command(name='trade', help='Trade a player from one team to another.\nUsage: !trade [from_team] [to_team] [player_name]')
async def trade(ctx, from_team: str, to_team: str, player_name: str):
    # Validate both teams exist
    if from_team not in TEAMS or to_team not in TEAMS:
        await ctx.send(f"One or both of the teams specified do not exist.")
        return

    # Validate the player is on the from_team
    if player_name not in TEAMS[from_team]['players']:
        await ctx.send(f"Player '{player_name}' is not on Team '{from_team}'.")
        return10

    # Perform the trade
    TEAMS[from_team]['players'].remove(player_name)
    TEAMS[to_team]['players'].append(player_name)
    save_data()

    # Confirm the trade to the user
    await ctx.send(f"Player '{player_name}' has been successfully traded from Team '{from_team}' to Team '{to_team}'.")




#############################################################
# Auction logics:
#############################################################

class AuctionView(discord.ui.View):
  def __init__(self, *, timeout=180):
      super().__init__(timeout=timeout)
      self.current_player = None
      self.current_bid = 0
      self.highest_bidder = None
      self.highest_bidder_id = None

  def reset_auction_state(self):
    # Reset all auction-related attributes to their default states
    self.current_bid = self.current_bid
    self.highest_bidder_id = None
    self.current_player = None  

  @discord.ui.button(label="Start Auction", style=discord.ButtonStyle.green, custom_id="start_auction")
  async def start_auction(self, interaction: discord.Interaction, button: discord.ui.Button):
    # Check if there are players for auction
    if not PLAYERS_FOR_AUCTION:
        await interaction.response.send_message("No players available for auction.", ephemeral=True)
        return

    # Select the first player from the list
    self.current_player = PLAYERS_FOR_AUCTION.pop(0)  # This removes the player from the auction list and sets them as the current player
    if self.current_bid < 100000:
       self.current_bid = self.current_player['base_price']
    formatted_bid = f"${self.current_bid:,.2f}"
    self.highest_bidder = None  # Reset highest bidder

    # Disable the 'Start Auction' button as the auction has now started
    button.disabled = True
    for item in self.children:  
        if item.custom_id != 'start_auction':
            item.disabled = False

    # Create a new AuctionView instance
    auction_view = AuctionView()


    # Construct the message to display player details
    player_details = f"**Current Player for Auction:** {self.current_player['name']} - {self.current_player['category']}\n"
    player_details += f"**Base Price:** {formatted_bid}\n"
    player_details += "Place your bids!"

    # Update the message with the new details
    await interaction.response.edit_message(content=player_details, view=self)

  @discord.ui.button(label="Next Player", style=discord.ButtonStyle.blurple, custom_id="next_player")
  async def next_player(self, interaction: discord.Interaction, button: discord.ui.Button):

      if str(interaction.user.id) not in AUTHORIZED_USER_IDS:
         # User is not authorized to click the Sold button
         await interaction.response.send_message("phir a gaya lode, abhi bhi tera kaam ni h.", ephemeral=True)
         return

      #  Move the current player to the 'unsold' list if they exist and were not sold
      if self.current_player and self.current_player not in UNSOLD_PLAYERS:
          UNSOLD_PLAYERS.append(self.current_player)
          save_data()  
    
      # Check if there is a next player in the auction list
      if not PLAYERS_FOR_AUCTION: 
          await interaction.response.edit_message(content="No more players left for auction.", view=None)
          return

      # Move to the next player in the list
      self.current_player = PLAYERS_FOR_AUCTION.pop(0)  # Get and remove the first player from the list
      self.current_bid = self.current_player['base_price']  # Set the starting bid to the player's base price
      self.highest_bidder = None  # Reset the highest bidder for the new player
      save_data()

      # Update the message to show the new player's details
      content = (
          f"Next up for auction: {self.current_player['name']} ({self.current_player['category']})\n"
          f"Starting bid: {self.current_bid}\n"
          f"Place your bids!"
      )

      # Disable the Next Player button if this is the last player
      if not PLAYERS_FOR_AUCTION:
          button.disabled = True

      # Respond to the interaction to update the message with new player details
      await interaction.response.edit_message(content=content, view=self)


  @discord.ui.button(label="Bid", style=discord.ButtonStyle.grey, custom_id="bid")
  async def bid(self, interaction: discord.Interaction, button: discord.ui.Button):
      # Check if there's already an auction in progress with a player
      if self.current_player is None:
          await interaction.response.send_message("No player is currently being auctioned!", ephemeral=True)
          return

      # Determine the increment based on the base price
      base_price = self.current_player.get('base_price', 0)  
      if base_price >= 200000:
          increment = 50000
      elif base_price >= 150000:
          increment = 30000
      elif base_price >= 100000:
          increment = 20000
      else:
          increment = 10000  

      # Check user's current purse against the proposed bid
      user_id_str = str(interaction.user.id) 
      if user_id_str in USERS and USERS[user_id_str]['purse'] >= self.current_bid + increment:
          # Increment the current bid since user has enough in their purse
          self.current_bid += increment
          formatted_bid = f"${self.current_bid:,.2f}"

          # Update the highest bidder to the user who clicked the button
          self.highest_bidder = interaction.user.display_name
          self.highest_bidder_id = interaction.user.id

          # Construct the new auction message
          auction_message = f"Current player: {self.current_player['name']}\nCurrent bid: {formatted_bid}\nHighest bidder: {self.highest_bidder}\n"

          # Update the message to reflect the new highest bid
          await interaction.response.edit_message(content=auction_message, view=self)
      else:
          # User does not have enough in their purse to make this bid
          await interaction.response.send_message("Lode ruk ja gareeb h tu", ephemeral=True)


  @discord.ui.button(label="Sold Player", style=discord.ButtonStyle.red, custom_id="sold_player")
  async def sold_player(self, interaction: discord.Interaction, button: discord.ui.Button):

      if str(interaction.user.id) not in AUTHORIZED_USER_IDS:
        # User is not authorized to click the Sold button
        await interaction.response.send_message("Chal bey lode tera kaam ni h.", ephemeral=True)
        return
        
      # Extract the winner's ID and validate against USERS
      if interaction.user.id == auctioneer_id or interaction.user.id == self.highest_bidder_id:
         winner_id = self.highest_bidder_id
         winner_id = str(winner_id)
      player_name = self.current_player["name"]
      team_name = None  
      print(f"Winner ID: {winner_id}")

      for team, details in TEAMS.items():
          print(f"Team: {team}, Owner ID: {details['owner']}")
          if str(details['owner']) == str(winner_id):  # Ensure both sides of the comparison are strings
              team_name = team
              print(f"Found matching team: {team}")  # Debug print to confirm matching team is found
              break

          if team_name is None:
             print("No matching team found for the winner.")





      if winner_id not in USERS:
          await interaction.response.send_message("The highest bidder is not registered.", ephemeral=True)
          await interaction.followup.send(f"{winner_id}")
          return

      # Check if the winner has enough in their purse
      winning_bid = self.current_bid
      if winner_id in USERS and USERS[winner_id]["purse"] >= winning_bid:
          # Deduct the winning bid from the winner's purse
          USERS[winner_id]["purse"] -= winning_bid

          # Transfer the player to the winning team
          if team_name in TEAMS and player_name not in TEAMS[team_name]["players"]:
              TEAMS[team_name]["players"].append(player_name)
              TEAMS[team_name]["purse"] -= winning_bid  
              save_data()

              # Send confirmation messages
              winning_bidder_user = await interaction.client.fetch_user(winner_id)
              await interaction.response.send_message(content=f"{player_name} has been sold to {team_name} for {winning_bid}.")
              await interaction.followup.send(f"{winning_bidder_user.display_name}'s new purse balance is {USERS[winner_id]['purse']}")

              if PLAYERS_FOR_AUCTION:  # Check if there are more players to auction
                self.current_player = PLAYERS_FOR_AUCTION.pop(0)  # Move to the next player
                self.current_bid = self.current_player['base_price']  # Reset the bid to the new player's base price
                self.highest_bidder_id = None  # Reset the highest bidder

                # Update the auction message for the new player
                formatted_bid = f"${self.current_bid:,.2f}"
                new_auction_message = f"Next up for auction: {self.current_player['name']}\nStarting bid: {self.current_bid}"
                await interaction.followup.send(new_auction_message)  # Notify channel of the new auction
              else:
                # No more players left to auction
                await interaction.followup.send("No more players left for auction.")
                self.reset_auction_state() 
          else:
              # Handle case where the team is at capacity or player is already in the team
              await interaction.followup.send(f"Cannot transfer {player_name} to {team_name}. Check team capacity or player membership.")
      else:
          # Handle case where the winner does not have enough funds
          await interaction.followup.send(f"{interaction.user.display_name} does not have enough in their purse to complete this purchase.")

@bot.command(name='start_auction')
@commands.has_permissions(administrator=True)
async def start_auction(ctx):
    view = AuctionView()
    await ctx.send("Auction is starting!", view=view)

bot.run(os.environ['TOKEN'])

